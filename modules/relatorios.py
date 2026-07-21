import streamlit as st
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import numpy as np
from pymongo import ASCENDING
import plotly.express as px
import plotly.graph_objects as go
from itertools import product
from calendar import month_name
import io
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== FUNÇÕES DE FILTRO DE DATA ====================
def _get_date_filters(collection, key_suffix=""):
    """
    Retorna datas inicial e final com base no filtro selecionado.
    key_suffix: sufixo para tornar a chave única quando chamado múltiplas vezes
    """
    # Primeiro, buscar os meses únicos com cadastros (para opção dinâmica)
    pipeline_meses = [
        { "$match": { "data_cadastro": { "$exists": True}}},
        { "$group": {
             "_id": {
                 "ano": { "$year":  "$data_cadastro"},
                 "mes": { "$month":  "$data_cadastro"}
            }
        }},
        { "$sort": { "_id.ano": -1,  "_id.mes": -1}}
    ]

    try:
        meses_disponiveis = list(collection.aggregate(pipeline_meses))
    except Exception:
        meses_disponiveis = []

    # Converter para lista legível
    opcoes_meses = []
    for item in meses_disponiveis:
        ano = item["_id"]["ano"]
        mes_num = item["_id"]["mes"]
        if 1 <= mes_num <= 12:
            mes_nome = month_name[mes_num]
            opcoes_meses.append({
                 "label": f"{mes_nome} {ano}",
                 "value": f"{ano}-{mes_num:02d}",
                 "inicio": datetime(ano, mes_num, 1),
                 "fim": datetime(ano, mes_num, 1) + relativedelta(months=1) - timedelta(days=1)
            })

    # Opções fixas
    agora = datetime.now()
    hoje_inicio = agora.replace(hour=0, minute=0, second=0, microsecond=0)
    hoje_fim = agora.replace(hour=23, minute=59, second=59, microsecond=999999)

    opcoes_fixas = {
         "Hoje": (hoje_inicio, hoje_fim),
         "Últimos 7 dias": (hoje_inicio - timedelta(days=6), hoje_fim),
         "Último mês": (
            (hoje_inicio.replace(day=1) - relativedelta(months=1)),
            hoje_fim
        ),
         "Últimos 3 meses": (
            hoje_inicio - relativedelta(months=3) + timedelta(days=1),
            hoje_fim
        ),
         "Últimos 6 meses": (
            hoje_inicio - relativedelta(months=6) + timedelta(days=1),
            hoje_fim
        ),
         "Último ano": (
            hoje_inicio.replace(month=1, day=1) - relativedelta(years=1),
            hoje_fim.replace(month=12, day=31)
        )
    }

    # Montar lista de opções para o selectbox
    todas_opcoes = list(opcoes_fixas.keys())
    if opcoes_meses:
        todas_opcoes.append("--- Meses Disponíveis ---")
        todas_opcoes.extend([m["label"] for m in opcoes_meses])

    # Seletor no topo - com key única
    escolha = st.selectbox(
         "Selecione o período:",
        todas_opcoes,
        index=0,
        key=f"periodo_selector_{key_suffix}" if key_suffix else "periodo_selector"
    )

    # Retornar datas conforme escolha
    if escolha in opcoes_fixas:
        return opcoes_fixas[escolha][0], opcoes_fixas[escolha][1]
    elif escolha in [m["label"] for m in opcoes_meses]:
        mes_selecionado = next(m for m in opcoes_meses if m["label"] == escolha)
        return mes_selecionado["inicio"], mes_selecionado["fim"]
    else:
        # Fallback: últimos 30 dias
        return hoje_inicio - timedelta(days=29), hoje_fim

# ==================== FUNÇÕES DE ANÁLISE ====================
def _taxa_conversao(collection, inicio, fim):
    """Calcula taxas de conversão geral e segmentada"""
    total_cadastros = collection.count_documents({
        "data_cadastro": {"$gte": inicio, "$lte": fim}
    })
    total_ativacoes = collection.count_documents({
         "data_cadastro": {"$gte": inicio, "$lte": fim},
         "seguiu_ativacao": "Sim"
    })

    restritivos = collection.count_documents({
         "data_cadastro": {"$gte": inicio, "$lte": fim},
         "restritivo": "Sim"
    })

    restritivos_ativados = collection.count_documents({
         "data_cadastro": {"$gte": inicio, "$lte": fim},
         "restritivo": "Sim",
         "seguiu_ativacao": "Sim"
    })

    nao_restritivos = total_cadastros - restritivos
    nao_restritivos_ativados = total_ativacoes - restritivos_ativados

    taxa_geral = (total_ativacoes / total_cadastros * 100) if total_cadastros > 0 else 0
    taxa_restritivos = (restritivos_ativados / restritivos * 100) if restritivos > 0 else 0
    taxa_nao_restritivos = (nao_restritivos_ativados / nao_restritivos * 100) if nao_restritivos > 0 else 0

    return {
         "total_cadastros": total_cadastros,
         "total_ativacoes": total_ativacoes,
         "taxa_geral": round(taxa_geral, 2),
         "restritivos": restritivos,
         "restritivos_ativados": restritivos_ativados,
         "taxa_restritivos": round(taxa_restritivos, 2),
         "nao_restritivos": nao_restritivos,
         "nao_restritivos_ativados": nao_restritivos_ativados,
         "taxa_nao_restritivos": round(taxa_nao_restritivos, 2)
    }

def _analise_restricoes_detalhada(collection, inicio, fim):
    """Análise detalhada de restrições: média, distribuição, etc."""
    pipeline = [
        { "$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "restritivo": "Sim",
            "seguiu_ativacao": "Sim",
            "restritivo_qtd_registros": { "$exists": True, "$ne": None}
        }},
        { "$group": {
            "_id": None,
            "total_ativacoes_com_restricao": { "$sum": 1},
            "quantidade_restricoes": { "$push": "$restritivo_qtd_registros"}
        }}
    ]
    result = list(collection.aggregate(pipeline))

    if result and result[0].get("quantidade_restricoes"):
        restricoes = [int(r) for r in result[0]["quantidade_restricoes"] if str(r).isdigit()]
        if restricoes:
            return {
                 "total_ativacoes_com_restricao": result[0]["total_ativacoes_com_restricao"],
                 "media_restricoes": round(np.mean(restricoes), 2),
                 "mediana_restricoes": round(np.median(restricoes), 2),
                 "max_restricoes": max(restricoes),
                 "min_restricoes": min(restricoes),
                 "desvio_padrao": round(np.std(restricoes), 2)
            }

    # Fallback se não houver dados detalhados
    total_ativacoes_com_restricao = collection.count_documents({
         "data_cadastro": {"$gte": inicio, "$lte": fim},
         "restritivo": "Sim",
         "seguiu_ativacao": "Sim"
    })

    return {
         "total_ativacoes_com_restricao": total_ativacoes_com_restricao,
         "media_restricoes": "N/A",
         "mediana_restricoes": "N/A",
         "max_restricoes": "N/A",
         "min_restricoes": "N/A",
         "desvio_padrao": "N/A"
    }

def _conversoes_por_periodo(collection, inicio, fim, agrupamento="dia"):
    """Retorna conversões por período (dia, semana, mês)"""
    # Buscar dados brutos primeiro
    pipeline = [
        { "$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim}
        }},
        { "$project": {
            "data_cadastro": 1,
            "seguiu_ativacao": 1
        }}
    ]
    result = list(collection.aggregate(pipeline))

    if not result:
        if agrupamento == "dia":
            return pd.DataFrame(columns=["Data", "Cadastros", "Ativações", "Taxa Conversão"])
        elif agrupamento == "semana":
            return pd.DataFrame(columns=["Semana", "Cadastros", "Ativações", "Taxa Conversão"])
        else:
            return pd.DataFrame(columns=["Mês", "Cadastros", "Ativações", "Taxa Conversão"])

    # Converter para DataFrame
    df = pd.DataFrame(result)
    df["data_cadastro"] = pd.to_datetime(df["data_cadastro"])

    # Criar coluna de agrupamento
    if agrupamento == "dia":
        df["periodo"] = df["data_cadastro"].dt.strftime("%Y-%m-%d")
        date_label = "Data"
    elif agrupamento == "semana":
        # Usar ano-semana (ex: 2026-04)
        df["periodo"] = df["data_cadastro"].dt.strftime("%Y-%W")
        date_label = "Semana"
    else:  # mes
        df["periodo"] = df["data_cadastro"].dt.strftime("%Y-%m")
        date_label = "Mês"

    # Agrupar no pandas
    df_agg = df.groupby(["periodo", "seguiu_ativacao"]).size().reset_index(name="total")

    # Pivotar
    df_pivot = df_agg.pivot(index="periodo", columns="seguiu_ativacao", values="total").fillna(0)

    # Garantir colunas "Sim" e "Não"
    if "Sim" not in df_pivot.columns:
        df_pivot["Sim"] = 0
    if "Não" not in df_pivot.columns:
        df_pivot["Não"] = 0

    df_pivot["Cadastros"] = df_pivot["Sim"] + df_pivot["Não"]
    df_pivot["Ativações"] = df_pivot["Sim"]
    df_pivot["Taxa Conversão"] = (df_pivot["Ativações"] / df_pivot["Cadastros"] * 100).round(2)

    df_pivot.reset_index(inplace=True)
    df_pivot.rename(columns={"periodo": date_label}, inplace=True)

    return df_pivot[[date_label, "Cadastros", "Ativações", "Taxa Conversão"]]

def _performance_origem(collection, inicio, fim):
    """Análise de performance por origem"""
    pipeline = [
        { "$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "origem": { "$ne": ""}
        }},
        { "$group": {
            "_id": "$origem",
            "total_cadastros": { "$sum": 1},
            "total_ativacoes": { "$sum": { "$cond": [{ "$eq": ["$seguiu_ativacao", "Sim"]}, 1, 0]}}
        }},
        { "$project": {
            "origem": "$_id",
            "total_cadastros": 1,
            "total_ativacoes": 1,
            "taxa_conversao": { "$round": [{ "$multiply": [{ "$divide": ["$total_ativacoes", "$total_cadastros"]}, 100]}, 2]}
        }},
        { "$sort": { "taxa_conversao": -1}}
    ]
    result = list(collection.aggregate(pipeline))
    return pd.DataFrame(result) if result else pd.DataFrame()

def _distribuicao_planos(collection, inicio, fim):
    """Distribuição de planos escolhidos"""
    pipeline = [
        { "$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "plano_escolhido": { "$ne": ""}
        }},
        { "$group": {
            "_id": "$plano_escolhido",
            "total": { "$sum": 1},
            "ativacoes": { "$sum": { "$cond": [{ "$eq": ["$seguiu_ativacao", "Sim"]}, 1, 0]}}
        }},
        { "$project": {
            "plano": "$_id",
            "total": 1,
            "ativacoes": 1,
            "taxa_conversao": { "$round": [{ "$multiply": [{ "$divide": ["$ativacoes", "$total"]}, 100]}, 2]}
        }},
        { "$sort": { "total": -1}}
    ]
    result = list(collection.aggregate(pipeline))
    return pd.DataFrame(result) if result else pd.DataFrame()

def _tempo_para_ativacao(collection, inicio, fim):
    """Análise do tempo médio para ativação"""
    # Busca clientes que ativaram
    ativados = list(collection.find(
        {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "seguiu_ativacao": "Sim",
            "data_ativacao": { "$exists": True}
        },
        {
            "data_cadastro": 1,
            "data_ativacao": 1
        }
    ))
    if not ativados:
        return {"media_dias": 0, "mediana_dias": 0, "total_ativados": 0}

    tempos = []
    for cliente in ativados:
        if cliente.get("data_ativacao") and cliente.get("data_cadastro"):
            if isinstance(cliente["data_ativacao"], datetime) and isinstance(cliente["data_cadastro"], datetime):
                dias = (cliente["data_ativacao"] - cliente["data_cadastro"]).days
                tempos.append(dias)

    if tempos:
        return {
             "media_dias": round(np.mean(tempos), 2),
             "mediana_dias": round(np.median(tempos), 2),
             "total_ativados": len(tempos),
             "max_dias": max(tempos),
             "min_dias": min(tempos)
        }

    return {"media_dias": 0, "mediana_dias": 0, "total_ativados": len(ativados)}

def _funil_conversao(collection, inicio, fim):
    """Funil de conversão completo"""
    total_cadastros = collection.count_documents({
        "data_cadastro": {"$gte": inicio, "$lte": fim}
    })
    restritivos = collection.count_documents({
         "data_cadastro": {"$gte": inicio, "$lte": fim},
         "restritivo": "Sim"
    })

    nao_restritivos = total_cadastros - restritivos

    ativacoes_total = collection.count_documents({
         "data_cadastro": {"$gte": inicio, "$lte": fim},
         "seguiu_ativacao": "Sim"
    })

    ativacoes_restritivos = collection.count_documents({
         "data_cadastro": {"$gte": inicio, "$lte": fim},
         "restritivo": "Sim",
         "seguiu_ativacao": "Sim"
    })

    ativacoes_nao_restritivos = ativacoes_total - ativacoes_restritivos

    return {
         "total_cadastros": total_cadastros,
         "restritivos": restritivos,
         "nao_restritivos": nao_restritivos,
         "ativacoes_total": ativacoes_total,
         "ativacoes_restritivos": ativacoes_restritivos,
         "ativacoes_nao_restritivos": ativacoes_nao_restritivos
    }

def _indicacoes_por_mes(collection, inicio, fim):
    """Retorna o número de clientes que seguiram para ativação por mês (indicados)."""
    pipeline = [
        { "$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "seguiu_ativacao": "Sim",
            "codigo_indicador": { "$ne": None, "$ne": ""}
        }},
        { "$group": {
            "_id": {
                "ano": {"$year": "$data_cadastro"},
                "mes": {"$month": "$data_cadastro"}
            },
            "total": {"$sum": 1}
        }},
        { "$sort": {"_id.ano": 1, "_id.mes": 1}}
    ]
    result = list(collection.aggregate(pipeline))
    df = pd.DataFrame(result)
    if df.empty:
        return pd.DataFrame(columns=["mes_ano", "total"])
    df["mes_ano"] = df["_id"].apply(lambda x: f"{x['ano']}-{x['mes']:02d}")
    df = df[["mes_ano", "total"]].sort_values("mes_ano")
    return df

def _top_indicadores(collection, inicio, fim, top_n=10):
    """Retorna os top N indicadores com base em quantos indicados seguiram para ativação."""
    pipeline = [
        { "$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "seguiu_ativacao": "Sim",
            "codigo_indicador": { "$ne": None, "$ne": ""}
        }},
        { "$group": {
            "_id": "$codigo_indicador",
            "total_indicados": {"$sum": 1}
        }},
        { "$sort": {"total_indicados": -1}},
        { "$limit": top_n}
    ]
    indicadores_codigo = list(collection.aggregate(pipeline))
    codigos = [item["_id"] for item in indicadores_codigo]
    if not codigos:
        return pd.DataFrame(columns=["Indicador", "Indicações Válidas"])

    indicadores_reais = list(collection.find(
        {"codigo_indicacao": {"$in": codigos}},
        {"nome_completo": 1, "codigo_indicacao": 1}
    ))

    mapa_codigo_nome = {
        ind.get("codigo_indicacao"): ind.get("nome_completo", "Nome não disponível")
        for ind in indicadores_reais
        if ind.get("codigo_indicacao") is not None
    }

    dados = []
    for item in indicadores_codigo:
        nome = mapa_codigo_nome.get(item["_id"], f"Código: {item['_id']}")
        dados.append({"Indicador": nome, "Indicações Válidas": item["total_indicados"]})

    return pd.DataFrame(dados)

def _detalhe_indicacoes(collection, inicio, fim):
    """Retorna lista detalhada de indicados que seguiram para ativação + nome do indicador."""
    indicados = list(collection.find(
        {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "seguiu_ativacao": "Sim",
            "codigo_indicador": { "$ne": None, "$ne": ""}
        },
        {
            "nome_completo": 1,
            "data_cadastro": 1,
            "codigo_indicador": 1
        }
    ))
    if not indicados:
        return pd.DataFrame()

    codigos_indicadores = list(set([
        ind.get("codigo_indicador") for ind in indicados
        if ind.get("codigo_indicador")
    ]))

    if not codigos_indicadores:
        return pd.DataFrame()

    indicadores = list(collection.find(
        {"codigo_indicacao": {"$in": codigos_indicadores}},
        {"codigo_indicacao": 1, "nome_completo": 1}
    ))

    mapa = {
        ind.get("codigo_indicacao"): ind.get("nome_completo", "Não encontrado")
        for ind in indicadores
        if ind.get("codigo_indicacao") is not None
    }

    linhas = []
    for ind in indicados:
        cod_ind = ind.get("codigo_indicador")
        if not cod_ind:
            continue
        linhas.append({
             "Indicado": ind.get("nome_completo", "Nome não informado"),
             "Data Cadastro": ind["data_cadastro"].strftime("%d/%m/%Y") if isinstance(ind.get("data_cadastro"), datetime) else str(ind.get("data_cadastro", " ")),
             "Código Indicador": cod_ind,
             "Quem Indicou": mapa.get(cod_ind, "Não encontrado")
        })

    df = pd.DataFrame(linhas)
    return df.sort_values("Data Cadastro", ascending=False)

def _cadastros_por_dia(collection, inicio, fim):
    pipeline = [
        { "$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim}
        }},
        { "$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$data_cadastro"}},
            "total": {"$sum": 1}
        }},
        { "$sort": {"_id": 1}}
    ]
    result = list(collection.aggregate(pipeline))
    df = pd.DataFrame(result)
    if df.empty:
        return pd.DataFrame(columns=["_id", "total"])
    df.columns = ["data", "cadastros"]
    df["data"] = pd.to_datetime(df["data"])
    return df.set_index("data").asfreq('D', fill_value=0).reset_index()

# ==================== FUNÇÕES PARA ANÁLISE DE NÃO CONVERSÃO ====================
def _analise_nao_conversao_detalhada(collection, inicio, fim):
    """Análise detalhada de não conversão usando o campo motivo_recusa_ativacao"""
    
    pipeline = [
        {"$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "seguiu_ativacao": "Não"
        }},
        {"$group": {
            "_id": "$motivo_recusa_ativacao",
            "total": {"$sum": 1},
            "clientes": {"$push": {
                "nome": "$nome_completo",
                "data": "$data_cadastro",
                "restritivo": "$restritivo",
                "detalhes": "$detalhes_recusa_ativacao"
            }}
        }},
        {"$sort": {"total": -1}}
    ]
    
    result = list(collection.aggregate(pipeline))
    
    restritivos = 0
    outros_motivos = 0
    detalhes_por_motivo = {}
    
    for item in result:
        motivo = item["_id"] if item["_id"] else "Não informado"
        total = item["total"]
        detalhes_por_motivo[motivo] = {
            "total": total,
            "clientes": item["clientes"]
        }
        
        if motivo == "Restritivos (SPC/Serasa)":
            restritivos += total
        else:
            outros_motivos += total
    
    total_nao_conversao = restritivos + outros_motivos
    
    return {
        "total_nao_conversao": total_nao_conversao,
        "restritivos": restritivos,
        "outros_motivos": outros_motivos,
        "detalhes_por_motivo": detalhes_por_motivo,
        "raw_data": result
    }

def _tabela_nao_conversao_detalhada(collection, inicio, fim):
    """Retorna DataFrame detalhado de não conversões"""
    
    docs = list(collection.find(
        {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "seguiu_ativacao": "Não"
        },
        {
            "nome_completo": 1,
            "celular": 1,
            "cpf": 1,
            "data_cadastro": 1,
            "motivo_recusa_ativacao": 1,
            "detalhes_recusa_ativacao": 1,
            "restritivo": 1,
            "restritivo_qtd_registros": 1,
            "origem": 1,
            "plano_escolhido": 1,
            "observacoes": 1
        }
    ).sort("data_cadastro", -1))
    
    if not docs:
        return pd.DataFrame()
    
    dados = []
    for doc in docs:
        motivo = doc.get("motivo_recusa_ativacao") or "Não informado"
        
        if motivo == "Restritivos (SPC/Serasa)":
            categoria = "Restritivo"
        else:
            categoria = "Outros Motivos"
        
        dados.append({
            "Nome": doc.get("nome_completo", ""),
            "Celular": doc.get("celular", ""),
            "CPF": doc.get("cpf", ""),
            "Data Cadastro": doc.get("data_cadastro").strftime("%d/%m/%Y") if isinstance(doc.get("data_cadastro"), datetime) else "",
            "Motivo Recusa": motivo,
            "Categoria": categoria,
            "É Restritivo?": "Sim" if doc.get("restritivo") == "Sim" else "Não",
            "Qtd Restrições": doc.get("restritivo_qtd_registros", ""),
            "Detalhes": doc.get("detalhes_recusa_ativacao", ""),
            "Origem": doc.get("origem", ""),
            "Plano": doc.get("plano_escolhido", ""),
            "Observações": doc.get("observacoes", "")
        })
    
    return pd.DataFrame(dados)

def _evolucao_motivos(collection, inicio, fim):
    """Evolução dos motivos ao longo do tempo (por mês)"""
    
    pipeline = [
        {"$match": {
            "data_cadastro": {"$gte": inicio, "$lte": fim},
            "seguiu_ativacao": "Não"
        }},
        {"$group": {
            "_id": {
                "ano": {"$year": "$data_cadastro"},
                "mes": {"$month": "$data_cadastro"},
                "motivo": "$motivo_recusa_ativacao"
            },
            "total": {"$sum": 1}
        }},
        {"$sort": {"_id.ano": 1, "_id.mes": 1}}
    ]
    
    result = list(collection.aggregate(pipeline))
    
    if not result:
        return pd.DataFrame()
    
    df = pd.DataFrame(result)
    df["mes_ano"] = df["_id"].apply(lambda x: f"{x['ano']}-{x['mes']:02d}")
    df["motivo"] = df["_id"].apply(lambda x: x.get("motivo") or "Não informado")
    
    return df

# ==================== FUNÇÃO: ANÁLISE DE LEADS POR VENDEDORA (FOLLOW-UP) ====================
def _leads_por_vendedora_followup(collection, inicio, fim, modo_delegacao_ativo=False, atendente_delegado=None):
    """
    Retorna análise de leads em follow-up agrupados por vendedora.
    Considera apenas clientes que NÃO seguiram para ativação e NÃO são restritivos.
    """
    
    query = {
        "seguiu_ativacao": {"$ne": "Sim"},
        "restritivo": {"$ne": "Sim"},
        "status_followup": {"$ne": "removido"}
    }
    
    if inicio and fim:
        query["data_cadastro"] = {"$gte": inicio, "$lte": fim}
    
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$cadastrado_por",
            "total_leads": {"$sum": 1},
            "com_data_definida": {
                "$sum": {
                    "$cond": [
                        {"$and": [
                            {"$ne": ["$retorno_agendado", ""]},
                            {"$ne": ["$retorno_agendado", None]},
                            {"$ne": ["$retorno_agendado", " "]}
                        ]},
                        1, 0
                    ]
                }
            },
            "sem_data_definida": {
                "$sum": {
                    "$cond": [
                        {"$or": [
                            {"$eq": ["$retorno_agendado", ""]},
                            {"$eq": ["$retorno_agendado", None]},
                            {"$eq": ["$retorno_agendado", " "]}
                        ]},
                        1, 0
                    ]
                }
            },
            "total_touches": {"$sum": "$touch_count"},
            "media_touches": {"$avg": "$touch_count"},
            "clientes": {"$push": {
                "nome": "$nome_completo",
                "celular": "$celular",
                "touch_count": "$touch_count",
                "retorno_agendado": "$retorno_agendado",
                "condominio_nome": "$condominio_nome",
                "bloco": "$bloco",
                "apartamento": "$apartamento",
                "observacoes_followup": "$observacoes_followup"  # <-- CAMPO CORRETO
            }}
        }},
        {"$project": {
            "vendedora": "$_id",
            "total_leads": 1,
            "com_data_definida": 1,
            "sem_data_definida": 1,
            "total_touches": 1,
            "media_touches": {"$ifNull": [{"$round": ["$media_touches", 2]}, 0]},
            "clientes": 1,
            "percentual_com_data": {
                "$cond": [
                    {"$eq": ["$total_leads", 0]},
                    0,
                    {"$round": [{"$multiply": [{"$divide": ["$com_data_definida", "$total_leads"]}, 100]}, 1]}
                ]
            }
        }},
        {"$sort": {"total_leads": -1}}
    ]
    
    result = list(collection.aggregate(pipeline))
    
    if modo_delegacao_ativo and atendente_delegado:
        if atendente_delegado == "Todos os atendentes":
            todos_atendentes = collection.distinct("cadastrado_por", {
                "seguiu_ativacao": {"$ne": "Sim"},
                "restritivo": {"$ne": "Sim"},
                "status_followup": {"$ne": "removido"}
            })
            todos_atendentes = [a for a in todos_atendentes if a]
        else:
            todos_atendentes = [atendente_delegado]
    else:
        todos_atendentes = collection.distinct("cadastrado_por", query)
        todos_atendentes = [a for a in todos_atendentes if a]
    
    resultado_dict = {r["_id"]: r for r in result if r["_id"]}
    resultado_final = []
    
    for atendente in sorted(todos_atendentes):
        if atendente in resultado_dict:
            resultado_final.append(resultado_dict[atendente])
        else:
            resultado_final.append({
                "vendedora": atendente,
                "total_leads": 0,
                "com_data_definida": 0,
                "sem_data_definida": 0,
                "total_touches": 0,
                "media_touches": 0,
                "clientes": [],
                "percentual_com_data": 0
            })
    
    return pd.DataFrame(resultado_final) if resultado_final else pd.DataFrame()

def render_relatorios_followup(clientes_collection):
    """Renderiza relatórios específicos de follow-up (apenas Admin/Diretoria)"""
    st.subheader("📊 Follow-up - Leads por Vendedora")
    st.markdown("Análise da distribuição de leads em follow-up por atendente/vendedora.")
    
    usuario_atual = st.session_state.get("nome_usuario", "")
    is_admin = (usuario_atual == "Diego Roberto")
    is_diretoria = (usuario_atual in ["Diretoria", "Gestão"])
    
    if not (is_admin or is_diretoria):
        st.error("⛔ Acesso restrito. Apenas Admin e Diretoria podem visualizar este relatório.")
        return
    
    try:
        from followup import get_modo_delegacao_estado
        estado_delegacao = get_modo_delegacao_estado(clientes_collection)
        modo_delegacao_ativo = estado_delegacao["ativo"]
        atendente_delegado = estado_delegacao["atendente"]
    except ImportError:
        modo_delegacao_ativo = False
        atendente_delegado = None
    
    # Usar sufixo para chave única
    inicio, fim = _get_date_filters(clientes_collection, key_suffix="followup")
    
    if modo_delegacao_ativo:
        if atendente_delegado == "Todos os atendentes":
            st.info("🔄 **Modo Delegação Ativo** - Exibindo dados de TODOS os atendentes")
        else:
            st.info(f"🔄 **Modo Delegação Ativo** - Exibindo dados apenas de **{atendente_delegado}**")
    
    df = _leads_por_vendedora_followup(
        clientes_collection, 
        inicio, 
        fim,
        modo_delegacao_ativo,
        atendente_delegado
    )
    
    if df.empty:
        st.warning("📭 Nenhum lead em follow-up encontrado no período selecionado.")
        return
    
    total_leads = df["total_leads"].sum()
    total_com_data = df["com_data_definida"].sum()
    total_sem_data = df["sem_data_definida"].sum()
    media_geral_touches = df["total_touches"].sum() / total_leads if total_leads > 0 else 0
    vendedoras_com_leads = len(df[df["total_leads"] > 0])
    
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total Leads em Follow-up", total_leads)
    col2.metric("Com Data Definida", total_com_data, f"{total_com_data/total_leads*100:.1f}%" if total_leads > 0 else "0%")
    col3.metric("Sem Data Definida", total_sem_data)
    col4.metric("Média Toques", f"{media_geral_touches:.2f}")
    col5.metric("Vendedoras Ativas", vendedoras_com_leads)
    
    st.markdown("---")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        df_ordenado = df.sort_values("total_leads", ascending=True)
        
        fig_barras = go.Figure()
        
        fig_barras.add_trace(go.Bar(
            y=df_ordenado["vendedora"],
            x=df_ordenado["total_leads"],
            name="Total Leads",
            marker_color="#3498db",
            text=df_ordenado["total_leads"],
            textposition="outside",
            orientation='h'
        ))
        
        fig_barras.add_trace(go.Bar(
            y=df_ordenado["vendedora"],
            x=df_ordenado["com_data_definida"],
            name="Com Data Definida",
            marker_color="#2ecc71",
            text=df_ordenado["com_data_definida"],
            textposition="outside",
            orientation='h'
        ))
        
        fig_barras.update_layout(
            title="Leads em Follow-up por Vendedora",
            barmode="group",
            height=max(400, len(df) * 35),
            xaxis_title="Quantidade",
            yaxis_title="",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        
        st.plotly_chart(fig_barras, use_container_width=True)
    
    with col2:
        df_pizza = df[df["total_leads"] > 0].copy()
        
        if not df_pizza.empty:
            if len(df_pizza) > 10:
                outros = df_pizza.iloc[10:]["total_leads"].sum()
                df_pizza = df_pizza.iloc[:10]
                if outros > 0:
                    df_pizza = pd.concat([
                        df_pizza,
                        pd.DataFrame([{"vendedora": "Outros", "total_leads": outros}])
                    ], ignore_index=True)
            
            fig_pizza = go.Figure(data=[go.Pie(
                labels=df_pizza["vendedora"],
                values=df_pizza["total_leads"],
                hole=0.3,
                textinfo='label+percent',
                textposition='auto'
            )])
            
            fig_pizza.update_layout(
                title="Distribuição de Leads",
                height=400
            )
            
            st.plotly_chart(fig_pizza, use_container_width=True)
    
    st.markdown("---")
    st.markdown("### 📋 Detalhamento por Vendedora")
    
    df_display = df[["vendedora", "total_leads", "com_data_definida", "sem_data_definida", 
                     "percentual_com_data", "media_touches", "total_touches"]].copy()
    
    df_display.columns = ["Vendedora", "Total Leads", "Com Data", "Sem Data", 
                          "% com Data", "Média Toques", "Total Toques"]
    
    df_display["% com Data"] = df_display["% com Data"].apply(lambda x: f"{x}%")
    
    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Vendedora": st.column_config.TextColumn("Vendedora", width="medium"),
            "Total Leads": st.column_config.NumberColumn("Total Leads", format="%d"),
            "Com Data": st.column_config.NumberColumn("Com Data", format="%d"),
            "Sem Data": st.column_config.NumberColumn("Sem Data", format="%d"),
            "% com Data": st.column_config.TextColumn("% com Data"),
            "Média Toques": st.column_config.NumberColumn("Média Toques", format="%.2f"),
            "Total Toques": st.column_config.NumberColumn("Total Toques", format="%d")
        }
    )
    
    st.markdown("---")
    st.markdown("### 📥 Exportar Dados")
    
    col1, col2 = st.columns(2)
    
    with col1:
        csv_data = df_display.to_csv(index=False, sep=';', encoding='utf-8-sig')
        st.download_button(
            label="📥 Exportar CSV",
            data=csv_data.encode('utf-8-sig'),
            file_name=f"followup_leads_por_vendedora_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    with col2:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_display.to_excel(writer, index=False, sheet_name='Leads por Vendedora')
            
            worksheet = writer.sheets['Leads por Vendedora']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        excel_data = output.getvalue()
        st.download_button(
            label="📊 Exportar Excel",
            data=excel_data,
            file_name=f"followup_leads_por_vendedora_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    st.markdown("---")
    st.markdown("### 👤 Detalhes por Vendedora")
    
    vendedora_selecionada = st.selectbox(
        "Selecione uma vendedora para ver detalhes:",
        options=sorted(df["vendedora"].unique())
    )
    
    if vendedora_selecionada:
        dados_vendedora = df[df["vendedora"] == vendedora_selecionada].iloc[0]
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Leads", dados_vendedora["total_leads"])
        col2.metric("Com Data Definida", dados_vendedora["com_data_definida"])
        col3.metric("Sem Data Definida", dados_vendedora["sem_data_definida"])
        
        media_valor = dados_vendedora.get("media_touches", 0)
        if pd.isna(media_valor) or media_valor is None:
            media_valor = 0
        col4.metric("Média Toques", f"{media_valor:.2f}")
        
        if dados_vendedora["total_leads"] > 0:
            with st.expander(f"📋 Ver leads de {vendedora_selecionada}"):
                clientes = dados_vendedora["clientes"]
                if clientes:
                    df_clientes = pd.DataFrame(clientes)
                    
                    # Mapear colunas existentes
                    colunas_esperadas = ["nome", "celular", "touch_count", "retorno_agendado", 
                                        "condominio_nome", "bloco", "apartamento", "observacoes_followup"]
                    colunas_existentes = [col for col in colunas_esperadas if col in df_clientes.columns]
                    
                    if not colunas_existentes:
                        st.info("Nenhum dado disponível para este vendedor.")
                        return
                    
                    df_clientes = df_clientes[colunas_existentes].copy()
                    
                    # Renomear colunas
                    mapa_renomeacao = {
                        "nome": "Nome",
                        "celular": "Celular",
                        "touch_count": "Toques",
                        "retorno_agendado": "Retorno",
                        "condominio_nome": "Condomínio",
                        "bloco": "Bloco",
                        "apartamento": "Apartamento",
                        "observacoes_followup": "Observação"
                    }
                    
                    for col_orig, col_dest in mapa_renomeacao.items():
                        if col_orig in df_clientes.columns:
                            df_clientes.rename(columns={col_orig: col_dest}, inplace=True)
                    
                    if "Retorno" in df_clientes.columns:
                        df_clientes["Retorno"] = df_clientes["Retorno"].apply(
                            lambda x: x if x and x != "" and x != " " else "Imediato"
                        )
                    
                    if "Observação" in df_clientes.columns:
                        df_clientes["Observação"] = df_clientes["Observação"].apply(
                            lambda x: x if x and x != "" and x != " " and pd.notna(x) else ""
                        )
                    
                    if "Bloco" in df_clientes.columns and "Apartamento" in df_clientes.columns:
                        df_clientes["Unidade"] = df_clientes.apply(
                            lambda row: f"{row['Bloco']} - {row['Apartamento']}" if row['Bloco'] and row['Apartamento'] 
                            else row['Bloco'] if row['Bloco'] 
                            else row['Apartamento'] if row['Apartamento'] 
                            else "",
                            axis=1
                        )
                    elif "Bloco" in df_clientes.columns:
                        df_clientes["Unidade"] = df_clientes["Bloco"]
                    elif "Apartamento" in df_clientes.columns:
                        df_clientes["Unidade"] = df_clientes["Apartamento"]
                    else:
                        df_clientes["Unidade"] = ""
                    
                    colunas_exibicao = ["Nome", "Celular", "Toques", "Retorno", "Condomínio", "Unidade", "Observação"]
                    colunas_existentes_exibicao = [col for col in colunas_exibicao if col in df_clientes.columns]
                    
                    if colunas_existentes_exibicao:
                        df_exibicao = df_clientes[colunas_existentes_exibicao].copy()
                        
                        column_config = {
                            "Nome": st.column_config.TextColumn("Nome", width="medium"),
                            "Celular": st.column_config.TextColumn("Celular", width="small"),
                            "Toques": st.column_config.NumberColumn("Toques", format="%d", width="small"),
                            "Retorno": st.column_config.TextColumn("Retorno", width="medium"),
                            "Condomínio": st.column_config.TextColumn("Condomínio", width="large"),
                            "Unidade": st.column_config.TextColumn("Unidade", width="small"),
                        }
                        
                        if "Observação" in df_exibicao.columns:
                            column_config["Observação"] = st.column_config.TextColumn("Observação", width="large")
                        
                        st.dataframe(
                            df_exibicao,
                            use_container_width=True,
                            hide_index=True,
                            column_config=column_config
                        )
                    else:
                        st.info("Nenhum dado disponível para exibição.")
        else:
            st.info(f"📭 {vendedora_selecionada} não possui leads em follow-up no período selecionado.")

# ==================== FUNÇÃO PRINCIPAL DE RENDERIZAÇÃO ====================
def render_relatorios(clientes_collection):
    st.header("📊 Dashboard de Relatórios - CRM")
    st.markdown("Visão analítica completa de leads, conversões, restrições, indicações e performance.")
    
    usuario_atual = st.session_state.get("nome_usuario", "")
    is_admin = (usuario_atual == "Diego Roberto")
    is_diretoria = (usuario_atual in ["Diretoria", "Gestão"])
    
    if is_admin or is_diretoria:
        tabs = st.tabs([
            "📊 Dashboard Geral",
            "📊 Follow-up Leads por Vendedora"
        ])
        
        with tabs[0]:
            # Dashboard Geral com sufixo vazio
            inicio, fim = _get_date_filters(clientes_collection, key_suffix="geral")
            st.caption(f"Período: **{inicio.strftime('%d/%m/%Y')}** a **{fim.strftime('%d/%m/%Y')}**")
            
            st.subheader("📈 KPIs Principais")
            conversao_data = _taxa_conversao(clientes_collection, inicio, fim)
            restricoes_data = _analise_restricoes_detalhada(clientes_collection, inicio, fim)
            tempo_data = _tempo_para_ativacao(clientes_collection, inicio, fim)

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Cadastros", conversao_data["total_cadastros"])
            col2.metric("Total Ativações", conversao_data["total_ativacoes"])
            col3.metric("Taxa Conversão", f"{conversao_data['taxa_geral']}%", 
                        delta=f"{conversao_data['taxa_nao_restritivos'] - conversao_data['taxa_restritivos']:.1f}% vs restritivos")
            col4.metric("Média Tempo Ativação", f"{tempo_data['media_dias']} dias")

            col5, col6, col7, col8 = st.columns(4)
            col5.metric("Com Restrição", conversao_data["restritivos"])
            col6.metric("Ativaram c/ Restrição", conversao_data["restritivos_ativados"])
            col7.metric("Taxa c/ Restrição", f"{conversao_data['taxa_restritivos']}%")
            col8.metric("Média Restrições", restricoes_data["media_restricoes"])

            st.markdown("---")
            st.subheader("🎯 Funil de Conversão")
            funil_data = _funil_conversao(clientes_collection, inicio, fim)

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Cadastros", funil_data["total_cadastros"])
                st.progress(100)
            with col2:
                taxa_restricao = (funil_data["restritivos"] / funil_data["total_cadastros"] * 100) if funil_data["total_cadastros"] > 0 else 0
                st.metric("Com Restrição", funil_data["restritivos"], f"{taxa_restricao:.1f}%")
                st.progress(int(taxa_restricao))
            with col3:
                taxa_conversao = (funil_data["ativacoes_total"] / funil_data["total_cadastros"] * 100) if funil_data["total_cadastros"] > 0 else 0
                st.metric("Total Ativações", funil_data["ativacoes_total"], f"{taxa_conversao:.1f}%")
                st.progress(int(taxa_conversao))

            fig_funil = go.Figure(go.Funnel(
                y = ["Cadastros", "Com Restrição", "Ativações"],
                x = [funil_data["total_cadastros"], funil_data["restritivos"], funil_data["ativacoes_total"]],
                textinfo = "value+percent initial"
            ))
            fig_funil.update_layout(height=300)
            st.plotly_chart(fig_funil, use_container_width=True)

            st.markdown("---")
            st.subheader("⚠️ Análise Detalhada de Restrições")
            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Ativações c/ Restrição", restricoes_data["total_ativacoes_com_restricao"])
            col2.metric("Média Restrições", restricoes_data["media_restricoes"])
            col3.metric("Mediana Restrições", restricoes_data["mediana_restricoes"])
            col4.metric("Máximo", restricoes_data["max_restricoes"])
            col5.metric("Mínimo", restricoes_data["min_restricoes"])
            if restricoes_data["desvio_padrao"] != "N/A":
                st.metric("Desvio Padrão", restricoes_data["desvio_padrao"])

            st.markdown("---")
            st.subheader("📊 Conversões por Período")
            col1, col2 = st.columns(2)
            with col1:
                agrupamento = st.selectbox("Agrupar por:", ["Dia", "Semana", "Mês"], index=2)
            with col2:
                mostrar_taxa = st.checkbox("Mostrar Taxa de Conversão", value=True)

            df_conversoes = _conversoes_por_periodo(clientes_collection, inicio, fim, agrupamento.lower())
            if not df_conversoes.empty:
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=df_conversoes[df_conversoes.columns[0]],
                    y=df_conversoes["Cadastros"],
                    name="Cadastros",
                    marker_color='lightblue'
                ))
                fig.add_trace(go.Bar(
                    x=df_conversoes[df_conversoes.columns[0]],
                    y=df_conversoes["Ativações"],
                    name="Ativações",
                    marker_color='green'
                ))
                if mostrar_taxa:
                    fig.add_trace(go.Scatter(
                        x=df_conversoes[df_conversoes.columns[0]],
                        y=df_conversoes["Taxa Conversão"],
                        name="Taxa Conversão (%)",
                        yaxis="y2",
                        mode='lines+markers',
                        line=dict(color='red', width=3),
                        marker=dict(size=8)
                    ))
                fig.update_layout(
                    barmode='group',
                    yaxis=dict(title="Quantidade"),
                    yaxis2=dict(title="Taxa Conversão (%)", overlaying="y", side="right"),
                    hovermode="x unified"
                )
                st.plotly_chart(fig, use_container_width=True)
                with st.expander("📋 Dados Detalhados"):
                    st.dataframe(df_conversoes, use_container_width=True)

            st.markdown("---")
            st.subheader("📍 Performance por Origem")
            df_origem = _performance_origem(clientes_collection, inicio, fim)
            if not df_origem.empty:
                col1, col2 = st.columns([2, 1])
                with col1:
                    fig_origem = px.bar(df_origem, x="origem", y=["total_cadastros", "total_ativacoes"], barmode="group", title="Cadastros vs Ativações por Origem")
                    st.plotly_chart(fig_origem, use_container_width=True)
                with col2:
                    fig_taxa = px.pie(df_origem, values="taxa_conversao", names="origem", title="Taxa de Conversão por Origem")
                    st.plotly_chart(fig_taxa, use_container_width=True)
                st.dataframe(df_origem.sort_values("taxa_conversao", ascending=False), use_container_width=True)

            st.markdown("---")
            st.subheader("💳 Distribuição de Planos")
            df_planos = _distribuicao_planos(clientes_collection, inicio, fim)
            if not df_planos.empty:
                col1, col2 = st.columns(2)
                with col1:
                    fig_planos = px.bar(df_planos, x="plano", y="total", color="ativacoes", title="Distribuição de Planos (Total vs Ativações)")
                    st.plotly_chart(fig_planos, use_container_width=True)
                with col2:
                    fig_conversao_plano = px.bar(df_planos, x="plano", y="taxa_conversao", title="Taxa de Conversão por Plano", color="taxa_conversao", color_continuous_scale="RdYlGn")
                    st.plotly_chart(fig_conversao_plano, use_container_width=True)

            st.markdown("---")
            st.subheader("🤝 Relatório de Indicações")
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("### 📅 Indicações que Resultaram em Ativação")
                df_indicacoes_mes = _indicacoes_por_mes(clientes_collection, inicio, fim)
                if not df_indicacoes_mes.empty:
                    df_indicacoes_mes["mes_ano"] = pd.to_datetime(df_indicacoes_mes["mes_ano"])
                    fig_indicacoes = px.bar(df_indicacoes_mes, x="mes_ano", y="total", title="Indicações com Ativação por Mês")
                    st.plotly_chart(fig_indicacoes, use_container_width=True)
            with col2:
                st.markdown("### 🏆 Top Indicadores")
                top_n = st.slider("Quantidade de indicadores no ranking", min_value=5, max_value=50, value=10, step=5)
                df_top_indicadores = _top_indicadores(clientes_collection, inicio, fim, top_n=top_n)
                if not df_top_indicadores.empty:
                    fig_top = px.bar(df_top_indicadores.head(10), x="Indicações Válidas", y="Indicador", orientation='h', title="Top 10 Indicadores")
                    st.plotly_chart(fig_top, use_container_width=True)

            st.markdown("### 📋 Indicados que Seguiram para Ativação")
            df_detalhe = _detalhe_indicacoes(clientes_collection, inicio, fim)
            if not df_detalhe.empty:
                st.dataframe(df_detalhe, use_container_width=True)
                csv = df_detalhe.to_csv(index=False).encode('utf-8')
                st.download_button(label="📥 Exportar Indicações Detalhadas (CSV)", data=csv, file_name="indicacoes_detalhadas.csv", mime="text/csv")

            st.markdown("---")
            st.subheader("🔍 Dados Brutos")
            with st.expander("📋 Visualizar Dados Brutos (últimos 200 cadastros)"):
                docs = list(clientes_collection.find(
                    {"data_cadastro": {"$gte": inicio, "$lte": fim}},
                    {"nome_completo": 1, "celular": 1, "data_cadastro": 1, "origem": 1, "restritivo": 1, "seguiu_ativacao": 1, "plano_escolhido": 1, "codigo_indicador": 1, "codigo_indicacao": 1, "restritivo_qtd_registros": 1, "data_ativacao": 1}
                ).sort("data_cadastro", -1).limit(200))
                for doc in docs:
                    if isinstance(doc.get("data_cadastro"), datetime):
                        doc["data_cadastro"] = doc["data_cadastro"].strftime("%d/%m/%Y %H:%M")
                    if isinstance(doc.get("data_ativacao"), datetime):
                        doc["data_ativacao"] = doc["data_ativacao"].strftime("%d/%m/%Y")
                if docs:
                    df_bruto = pd.DataFrame(docs)
                    st.dataframe(df_bruto, use_container_width=True)
                    csv_bruto = df_bruto.to_csv(index=False).encode('utf-8')
                    st.download_button(label="📥 Exportar Dados Brutos (CSV)", data=csv_bruto, file_name="dados_brutos.csv", mime="text/csv")

            st.markdown("---")
            st.subheader("⚠️ Análise de Risco - Exportação")
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown("Exporte dados dos clientes que **seguiram para ativação** para análise de risco e crédito.")
            with col2:
                docs_risco = list(clientes_collection.find(
                    {"data_cadastro": {"$gte": inicio, "$lte": fim}, "seguiu_ativacao": "Sim"},
                    {"nome_completo": 1, "cpf": 1, "celular": 1, "celular_contato_1": 1, "celular_contato_2": 1, "email": 1, "data_cadastro": 1, "data_ativacao": 1, "endereco": 1, "numero": 1, "complemento": 1, "bairro": 1, "cidade": 1, "restritivo": 1, "restritivo_qtd_registros": 1, "restritivo_ano_recente": 1, "restritivo_servico_internet": 1, "plano_escolhido": 1, "profissao": 1, "data_vencimento": 1, "origem": 1, "codigo_indicacao": 1, "cadastrado_por": 1}
                ).sort("data_cadastro", -1))
                if docs_risco:
                    dados_risco = []
                    for doc in docs_risco:
                        dados_risco.append({
                            "Nome Completo": doc.get("nome_completo", " "),
                            "CPF": doc.get("cpf", " "),
                            "Celular Principal": doc.get("celular", " "),
                            "Celular Contato 1": doc.get("celular_contato_1", " "),
                            "Celular Contato 2": doc.get("celular_contato_2", " "),
                            "Email": doc.get("email", " "),
                            "Data Cadastro": doc.get("data_cadastro").strftime("%d/%m/%Y %H:%M") if isinstance(doc.get("data_cadastro"), datetime) else " ",
                            "Data Ativacao": doc.get("data_ativacao").strftime("%d/%m/%Y") if isinstance(doc.get("data_ativacao"), datetime) else " ",
                            "Endereco": doc.get("endereco", " "),
                            "Numero": doc.get("numero", " "),
                            "Complemento": doc.get("complemento", " "),
                            "Bairro": doc.get("bairro", " "),
                            "Cidade": doc.get("cidade", " "),
                            "Restritivo": doc.get("restritivo", " "),
                            "Qtd Restricoes": doc.get("restritivo_qtd_registros", " "),
                            "Ano Restricao Mais Recente": doc.get("restritivo_ano_recente", " "),
                            "Possui Servico Internet": doc.get("restritivo_servico_internet", " "),
                            "Plano Escolhido": doc.get("plano_escolhido", " "),
                            "Profissao": doc.get("profissao", " "),
                            "Data Vencimento": doc.get("data_vencimento", " "),
                            "Origem": doc.get("origem", " "),
                            "Codigo Indicacao": doc.get("codigo_indicacao", " "),
                            "Cadastrado Por": doc.get("cadastrado_por", " ")
                        })
                    df_risco = pd.DataFrame(dados_risco)
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_risco.to_excel(writer, index=False, sheet_name='Analise de Risco')
                        worksheet = writer.sheets['Analise de Risco']
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    excel_data = output.getvalue()
                    st.download_button(label=f"📊 Exportar Análise de Risco ({len(docs_risco)} ativações)", data=excel_data, file_name=f"analise_risco_{inicio.strftime('%Y%m%d')}_{fim.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="export_risco_excel")
                else:
                    st.info("ℹ️ Nenhuma ativação no período selecionado")

            st.markdown("---")
            st.subheader("📈 Evolutivo Mensal de Conversão")
            pipeline_mensal = [
                {"$match": {"data_cadastro": {"$gte": inicio.replace(day=1) - relativedelta(months=11), "$lte": fim}}},
                {"$group": {
                    "_id": {"ano": {"$year": "$data_cadastro"}, "mes": {"$month": "$data_cadastro"}},
                    "total_cadastros": {"$sum": 1},
                    "total_ativacoes": {"$sum": {"$cond": [{"$eq": ["$seguiu_ativacao", "Sim"]}, 1, 0]}},
                    "nao_ativados_restritivos": {"$sum": {"$cond": [{"$and": [{"$eq": ["$seguiu_ativacao", "Não"]}, {"$eq": ["$restritivo", "Sim"]}]}, 1, 0]}},
                    "nao_ativados_outros": {"$sum": {"$cond": [{"$and": [{"$eq": ["$seguiu_ativacao", "Não"]}, {"$ne": ["$restritivo", "Sim"]}]}, 1, 0]}}
                }},
                {"$sort": {"_id.ano": 1, "_id.mes": 1}}
            ]
            dados_mensais = list(clientes_collection.aggregate(pipeline_mensal))
            if dados_mensais:
                df_mensal = pd.DataFrame(dados_mensais)
                df_mensal["mes_ano"] = df_mensal["_id"].apply(lambda x: f"{x['ano']}-{x['mes']:02d}")
                df_mensal["taxa_conversao"] = (df_mensal["total_ativacoes"] / df_mensal["total_cadastros"] * 100).round(2)
                fig_evolucao = go.Figure()
                fig_evolucao.add_trace(go.Bar(x=df_mensal["mes_ano"], y=df_mensal["total_cadastros"], name="Total Cadastros", marker_color="#3498db", text=df_mensal["total_cadastros"], textposition="outside"))
                fig_evolucao.add_trace(go.Bar(x=df_mensal["mes_ano"], y=df_mensal["total_ativacoes"], name="Total Ativações", marker_color="#2ecc71", text=df_mensal["total_ativacoes"], textposition="outside"))
                fig_evolucao.add_trace(go.Scatter(x=df_mensal["mes_ano"], y=df_mensal["taxa_conversao"], name="Taxa Conversão (%)", mode="lines+markers+text", line=dict(color="#e74c3c", width=3), marker=dict(size=8), text=df_mensal["taxa_conversao"].apply(lambda x: f"{x}%"), textposition="top center", yaxis="y2"))
                fig_evolucao.update_layout(title="Evolução Mensal: Cadastros, Ativações e Taxa de Conversão", barmode="group", xaxis=dict(title="Mês/Ano"), yaxis=dict(title="Quantidade", side="left"), yaxis2=dict(title="Taxa Conversão (%)", overlaying="y", side="right", showgrid=False), legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1), height=500, hovermode="x unified")
                st.plotly_chart(fig_evolucao, use_container_width=True)
                with st.expander("📋 Ver Dados Detalhados"):
                    df_display = df_mensal[["mes_ano", "total_cadastros", "total_ativacoes", "taxa_conversao"]].copy()
                    df_display.columns = ["Mês/Ano", "Total Cadastros", "Total Ativações", "Taxa Conversão (%)"]
                    st.dataframe(df_display, use_container_width=True)

            st.markdown("---")
            st.subheader("📊 Análise de Não Conversão")
            analise_data = _analise_nao_conversao_detalhada(clientes_collection, inicio, fim)
            df_detalhado = _tabela_nao_conversao_detalhada(clientes_collection, inicio, fim)
            if analise_data["total_nao_conversao"] > 0:
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Não Convertidos", analise_data["total_nao_conversao"])
                with col2:
                    perc_restritivos = (analise_data["restritivos"] / analise_data["total_nao_conversao"] * 100) if analise_data["total_nao_conversao"] > 0 else 0
                    st.metric("Por Restrição", analise_data["restritivos"], f"{perc_restritivos:.1f}%")
                with col3:
                    perc_outros = (analise_data["outros_motivos"] / analise_data["total_nao_conversao"] * 100) if analise_data["total_nao_conversao"] > 0 else 0
                    st.metric("Outros Motivos", analise_data["outros_motivos"], f"{perc_outros:.1f}%")
                st.markdown("---")
                col1, col2 = st.columns(2)
                with col1:
                    motivos = list(analise_data["detalhes_por_motivo"].keys())
                    valores = [analise_data["detalhes_por_motivo"][m]["total"] for m in motivos]
                    fig_pizza = go.Figure(data=[go.Pie(labels=motivos, values=valores, hole=0.3, textinfo='label+percent', textposition='outside', automargin=True)])
                    fig_pizza.update_layout(title="Distribuição por Motivo de Recusa", height=400, showlegend=True, legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1))
                    st.plotly_chart(fig_pizza, use_container_width=True)
                with col2:
                    df_comp = pd.DataFrame({"Categoria": ["Restritivos", "Outros Motivos"], "Quantidade": [analise_data["restritivos"], analise_data["outros_motivos"]]})
                    fig_barras = px.bar(df_comp, x="Categoria", y="Quantidade", color="Categoria", color_discrete_map={"Restritivos": "#e74c3c", "Outros Motivos": "#f39c12"}, text="Quantidade")
                    fig_barras.update_layout(title="Comparativo: Restritivos vs Outros Motivos", height=400, showlegend=False)
                    st.plotly_chart(fig_barras, use_container_width=True)
                st.markdown("### 📊 Top Motivos de Recusa")
                df_motivos = pd.DataFrame([{"Motivo": motivo, "Total": data["total"]} for motivo, data in analise_data["detalhes_por_motivo"].items()]).sort_values("Total", ascending=True)
                if not df_motivos.empty:
                    fig_top = px.bar(df_motivos, y="Motivo", x="Total", orientation='h', color="Total", color_continuous_scale="Reds", text="Total")
                    fig_top.update_layout(height=max(300, len(df_motivos) * 40), showlegend=False, xaxis_title="Quantidade", yaxis_title="")
                    st.plotly_chart(fig_top, use_container_width=True)
                st.markdown("### 💡 Insights")
                col1, col2 = st.columns(2)
                with col1:
                    if analise_data["restritivos"] > analise_data["outros_motivos"]:
                        st.warning(f"⚠️ **Maioria por restrição**: {perc_restritivos:.1f}% das não conversões são por restrições financeiras")
                    else:
                        st.info(f"ℹ️ **Maioria por outros motivos**: {perc_outros:.1f}% das não conversões são por outros fatores")
                with col2:
                    motivo_top = df_motivos.iloc[-1]["Motivo"] if not df_motivos.empty else "N/A"
                    total_top = df_motivos.iloc[-1]["Total"] if not df_motivos.empty else 0
                    st.metric("Motivo Mais Comum", motivo_top, f"{total_top} ocorrências")
                st.markdown("---")
                st.markdown("### 📋 Detalhamento das Não Conversões")
                if not df_detalhado.empty:
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        filtro_categoria = st.multiselect("Filtrar por Categoria", options=df_detalhado["Categoria"].unique(), default=df_detalhado["Categoria"].unique())
                    with col2:
                        filtro_motivo = st.multiselect("Filtrar por Motivo", options=df_detalhado["Motivo Recusa"].unique(), default=df_detalhado["Motivo Recusa"].unique())
                    with col3:
                        filtro_restritivo = st.selectbox("É Restritivo?", options=["Todos", "Sim", "Não"], index=0)
                    df_filtrado = df_detalhado.copy()
                    if filtro_categoria:
                        df_filtrado = df_filtrado[df_filtrado["Categoria"].isin(filtro_categoria)]
                    if filtro_motivo:
                        df_filtrado = df_filtrado[df_filtrado["Motivo Recusa"].isin(filtro_motivo)]
                    if filtro_restritivo != "Todos":
                        df_filtrado = df_filtrado[df_filtrado["É Restritivo?"] == filtro_restritivo]
                    st.dataframe(df_filtrado, use_container_width=True, hide_index=True)
                    st.markdown("#### 📥 Exportar Dados")
                    col1, col2 = st.columns(2)
                    with col1:
                        csv = df_filtrado.to_csv(index=False, sep=';').encode('utf-8')
                        st.download_button(label="📥 Exportar CSV", data=csv, file_name=f"nao_conversao_{inicio.strftime('%Y%m%d')}_{fim.strftime('%Y%m%d')}.csv", mime="text/csv")
                    with col2:
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_filtrado.to_excel(writer, index=False, sheet_name='Não Conversão')
                            worksheet = writer.sheets['Não Conversão']
                            header_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
                            header_font = Font(color="FFFFFF", bold=True)
                            for cell in worksheet[1]:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                worksheet.column_dimensions[column_letter].width = adjusted_width
                        excel_data = output.getvalue()
                        st.download_button(label="📊 Exportar Excel", data=excel_data, file_name=f"nao_conversao_{inicio.strftime('%Y%m%d')}_{fim.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.markdown("---")
                st.markdown("### 📈 Evolução Temporal por Motivo")
                df_evolucao = _evolucao_motivos(clientes_collection, inicio, fim)
                if not df_evolucao.empty:
                    df_pivot = df_evolucao.pivot_table(index="mes_ano", columns="motivo", values="total", fill_value=0).reset_index()
                    df_melt = df_pivot.melt(id_vars=["mes_ano"], var_name="Motivo", value_name="Quantidade")
                    fig_evolucao = px.line(df_melt, x="mes_ano", y="Quantidade", color="Motivo", markers=True, title="Evolução Mensal dos Motivos de Recusa")
                    fig_evolucao.update_layout(height=400, xaxis_title="Mês/Ano", yaxis_title="Quantidade", hovermode="x unified")
                    st.plotly_chart(fig_evolucao, use_container_width=True)
            else:
                st.success("✅ Nenhuma não conversão no período selecionado (taxa de conversão 100%!)")

            st.markdown("---")
            st.subheader("📋 Resumo Executivo")
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("### 🎯 Principais Insights")
                st.markdown(f"""
                - **Total de cadastros:** {conversao_data['total_cadastros']}
                - **Taxa de conversão geral:** {conversao_data['taxa_geral']}%
                - **Clientes com restrição:** {conversao_data['restritivos']} ({conversao_data['taxa_restritivos']}% converteram)
                - **Média de tempo para ativação:** {tempo_data['media_dias']} dias
                - **Indicações que ativaram:** {df_indicacoes_mes['total'].sum() if not df_indicacoes_mes.empty else 0}
                """)
            with col2:
                st.markdown("### ⚠️ Pontos de Atenção")
                if conversao_data['taxa_restritivos'] < conversao_data['taxa_nao_restritivos']:
                    st.warning(f"Taxa de conversão com restrição ({conversao_data['taxa_restritivos']}%) é menor que sem restrição ({conversao_data['taxa_nao_restritivos']})")
                if tempo_data['media_dias'] > 7:
                    st.warning(f"Tempo médio para ativação ({tempo_data['media_dias']} dias) está acima de 7 dias")
                if restricoes_data['media_restricoes'] != "N/A" and restricoes_data['media_restricoes'] > 2:
                    st.warning(f"Média de restrições por cliente ({restricoes_data['media_restricoes']}) está elevada")
        
        with tabs[1]:
            render_relatorios_followup(clientes_collection)
    
    else:
        # Usuário normal vê apenas o dashboard geral
        inicio, fim = _get_date_filters(clientes_collection, key_suffix="geral")
        st.caption(f"Período: **{inicio.strftime('%d/%m/%Y')}** a **{fim.strftime('%d/%m/%Y')}**")
        
        st.subheader("📈 KPIs Principais")
        conversao_data = _taxa_conversao(clientes_collection, inicio, fim)
        restricoes_data = _analise_restricoes_detalhada(clientes_collection, inicio, fim)
        tempo_data = _tempo_para_ativacao(clientes_collection, inicio, fim)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Cadastros", conversao_data["total_cadastros"])
        col2.metric("Total Ativações", conversao_data["total_ativacoes"])
        col3.metric("Taxa Conversão", f"{conversao_data['taxa_geral']}%", 
                    delta=f"{conversao_data['taxa_nao_restritivos'] - conversao_data['taxa_restritivos']:.1f}% vs restritivos")
        col4.metric("Média Tempo Ativação", f"{tempo_data['media_dias']} dias")

        col5, col6, col7, col8 = st.columns(4)
        col5.metric("Com Restrição", conversao_data["restritivos"])
        col6.metric("Ativaram c/ Restrição", conversao_data["restritivos_ativados"])
        col7.metric("Taxa c/ Restrição", f"{conversao_data['taxa_restritivos']}%")
        col8.metric("Média Restrições", restricoes_data["media_restricoes"])

        st.markdown("---")
        st.subheader("🎯 Funil de Conversão")
        funil_data = _funil_conversao(clientes_collection, inicio, fim)

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Cadastros", funil_data["total_cadastros"])
            st.progress(100)
        with col2:
            taxa_restricao = (funil_data["restritivos"] / funil_data["total_cadastros"] * 100) if funil_data["total_cadastros"] > 0 else 0
            st.metric("Com Restrição", funil_data["restritivos"], f"{taxa_restricao:.1f}%")
            st.progress(int(taxa_restricao))
        with col3:
            taxa_conversao = (funil_data["ativacoes_total"] / funil_data["total_cadastros"] * 100) if funil_data["total_cadastros"] > 0 else 0
            st.metric("Total Ativações", funil_data["ativacoes_total"], f"{taxa_conversao:.1f}%")
            st.progress(int(taxa_conversao))

        fig_funil = go.Figure(go.Funnel(
            y = ["Cadastros", "Com Restrição", "Ativações"],
            x = [funil_data["total_cadastros"], funil_data["restritivos"], funil_data["ativacoes_total"]],
            textinfo = "value+percent initial"
        ))
        fig_funil.update_layout(height=300)
        st.plotly_chart(fig_funil, use_container_width=True)

        st.markdown("---")
        st.subheader("⚠️ Análise Detalhada de Restrições")
        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("Ativações c/ Restrição", restricoes_data["total_ativacoes_com_restricao"])
        col2.metric("Média Restrições", restricoes_data["media_restricoes"])
        col3.metric("Mediana Restrições", restricoes_data["mediana_restricoes"])
        col4.metric("Máximo", restricoes_data["max_restricoes"])
        col5.metric("Mínimo", restricoes_data["min_restricoes"])
        if restricoes_data["desvio_padrao"] != "N/A":
            st.metric("Desvio Padrão", restricoes_data["desvio_padrao"])

        st.markdown("---")
        st.subheader("📊 Conversões por Período")
        col1, col2 = st.columns(2)
        with col1:
            agrupamento = st.selectbox("Agrupar por:", ["Dia", "Semana", "Mês"], index=2)
        with col2:
            mostrar_taxa = st.checkbox("Mostrar Taxa de Conversão", value=True)

        df_conversoes = _conversoes_por_periodo(clientes_collection, inicio, fim, agrupamento.lower())
        if not df_conversoes.empty:
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_conversoes[df_conversoes.columns[0]], y=df_conversoes["Cadastros"], name="Cadastros", marker_color='lightblue'))
            fig.add_trace(go.Bar(x=df_conversoes[df_conversoes.columns[0]], y=df_conversoes["Ativações"], name="Ativações", marker_color='green'))
            if mostrar_taxa:
                fig.add_trace(go.Scatter(x=df_conversoes[df_conversoes.columns[0]], y=df_conversoes["Taxa Conversão"], name="Taxa Conversão (%)", yaxis="y2", mode='lines+markers', line=dict(color='red', width=3), marker=dict(size=8)))
            fig.update_layout(barmode='group', yaxis=dict(title="Quantidade"), yaxis2=dict(title="Taxa Conversão (%)", overlaying="y", side="right"), hovermode="x unified")
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("📋 Dados Detalhados"):
                st.dataframe(df_conversoes, use_container_width=True)

        st.markdown("---")
        st.subheader("📍 Performance por Origem")
        df_origem = _performance_origem(clientes_collection, inicio, fim)
        if not df_origem.empty:
            col1, col2 = st.columns([2, 1])
            with col1:
                fig_origem = px.bar(df_origem, x="origem", y=["total_cadastros", "total_ativacoes"], barmode="group", title="Cadastros vs Ativações por Origem")
                st.plotly_chart(fig_origem, use_container_width=True)
            with col2:
                fig_taxa = px.pie(df_origem, values="taxa_conversao", names="origem", title="Taxa de Conversão por Origem")
                st.plotly_chart(fig_taxa, use_container_width=True)
            st.dataframe(df_origem.sort_values("taxa_conversao", ascending=False), use_container_width=True)

        st.markdown("---")
        st.subheader("💳 Distribuição de Planos")
        df_planos = _distribuicao_planos(clientes_collection, inicio, fim)
        if not df_planos.empty:
            col1, col2 = st.columns(2)
            with col1:
                fig_planos = px.bar(df_planos, x="plano", y="total", color="ativacoes", title="Distribuição de Planos (Total vs Ativações)")
                st.plotly_chart(fig_planos, use_container_width=True)
            with col2:
                fig_conversao_plano = px.bar(df_planos, x="plano", y="taxa_conversao", title="Taxa de Conversão por Plano", color="taxa_conversao", color_continuous_scale="RdYlGn")
                st.plotly_chart(fig_conversao_plano, use_container_width=True)

        st.markdown("---")
        st.subheader("🤝 Relatório de Indicações")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("### 📅 Indicações que Resultaram em Ativação")
            df_indicacoes_mes = _indicacoes_por_mes(clientes_collection, inicio, fim)
            if not df_indicacoes_mes.empty:
                df_indicacoes_mes["mes_ano"] = pd.to_datetime(df_indicacoes_mes["mes_ano"])
                fig_indicacoes = px.bar(df_indicacoes_mes, x="mes_ano", y="total", title="Indicações com Ativação por Mês")
                st.plotly_chart(fig_indicacoes, use_container_width=True)
        with col2:
            st.markdown("### 🏆 Top Indicadores")
            top_n = st.slider("Quantidade de indicadores no ranking", min_value=5, max_value=50, value=10, step=5)
            df_top_indicadores = _top_indicadores(clientes_collection, inicio, fim, top_n=top_n)
            if not df_top_indicadores.empty:
                fig_top = px.bar(df_top_indicadores.head(10), x="Indicações Válidas", y="Indicador", orientation='h', title="Top 10 Indicadores")
                st.plotly_chart(fig_top, use_container_width=True)

        st.markdown("### 📋 Indicados que Seguiram para Ativação")
        df_detalhe = _detalhe_indicacoes(clientes_collection, inicio, fim)
        if not df_detalhe.empty:
            st.dataframe(df_detalhe, use_container_width=True)
            csv = df_detalhe.to_csv(index=False).encode('utf-8')
            st.download_button(label="📥 Exportar Indicações Detalhadas (CSV)", data=csv, file_name="indicacoes_detalhadas.csv", mime="text/csv")

        st.markdown("---")
        st.subheader("🔍 Dados Brutos")
        with st.expander("📋 Visualizar Dados Brutos (últimos 200 cadastros)"):
            docs = list(clientes_collection.find(
                {"data_cadastro": {"$gte": inicio, "$lte": fim}},
                {"nome_completo": 1, "celular": 1, "data_cadastro": 1, "origem": 1, "restritivo": 1, "seguiu_ativacao": 1, "plano_escolhido": 1, "codigo_indicador": 1, "codigo_indicacao": 1, "restritivo_qtd_registros": 1, "data_ativacao": 1}
            ).sort("data_cadastro", -1).limit(200))
            for doc in docs:
                if isinstance(doc.get("data_cadastro"), datetime):
                    doc["data_cadastro"] = doc["data_cadastro"].strftime("%d/%m/%Y %H:%M")
                if isinstance(doc.get("data_ativacao"), datetime):
                    doc["data_ativacao"] = doc["data_ativacao"].strftime("%d/%m/%Y")
            if docs:
                df_bruto = pd.DataFrame(docs)
                st.dataframe(df_bruto, use_container_width=True)
                csv_bruto = df_bruto.to_csv(index=False).encode('utf-8')
                st.download_button(label="📥 Exportar Dados Brutos (CSV)", data=csv_bruto, file_name="dados_brutos.csv", mime="text/csv")

        st.markdown("---")
        st.subheader("⚠️ Análise de Risco - Exportação")
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("Exporte dados dos clientes que **seguiram para ativação** para análise de risco e crédito.")
        with col2:
            docs_risco = list(clientes_collection.find(
                {"data_cadastro": {"$gte": inicio, "$lte": fim}, "seguiu_ativacao": "Sim"},
                {"nome_completo": 1, "cpf": 1, "celular": 1, "celular_contato_1": 1, "celular_contato_2": 1, "email": 1, "data_cadastro": 1, "data_ativacao": 1, "endereco": 1, "numero": 1, "complemento": 1, "bairro": 1, "cidade": 1, "restritivo": 1, "restritivo_qtd_registros": 1, "restritivo_ano_recente": 1, "restritivo_servico_internet": 1, "plano_escolhido": 1, "profissao": 1, "data_vencimento": 1, "origem": 1, "codigo_indicacao": 1, "cadastrado_por": 1}
            ).sort("data_cadastro", -1))
            if docs_risco:
                dados_risco = []
                for doc in docs_risco:
                    dados_risco.append({
                        "Nome Completo": doc.get("nome_completo", " "),
                        "CPF": doc.get("cpf", " "),
                        "Celular Principal": doc.get("celular", " "),
                        "Celular Contato 1": doc.get("celular_contato_1", " "),
                        "Celular Contato 2": doc.get("celular_contato_2", " "),
                        "Email": doc.get("email", " "),
                        "Data Cadastro": doc.get("data_cadastro").strftime("%d/%m/%Y %H:%M") if isinstance(doc.get("data_cadastro"), datetime) else " ",
                        "Data Ativacao": doc.get("data_ativacao").strftime("%d/%m/%Y") if isinstance(doc.get("data_ativacao"), datetime) else " ",
                        "Endereco": doc.get("endereco", " "),
                        "Numero": doc.get("numero", " "),
                        "Complemento": doc.get("complemento", " "),
                        "Bairro": doc.get("bairro", " "),
                        "Cidade": doc.get("cidade", " "),
                        "Restritivo": doc.get("restritivo", " "),
                        "Qtd Restricoes": doc.get("restritivo_qtd_registros", " "),
                        "Ano Restricao Mais Recente": doc.get("restritivo_ano_recente", " "),
                        "Possui Servico Internet": doc.get("restritivo_servico_internet", " "),
                        "Plano Escolhido": doc.get("plano_escolhido", " "),
                        "Profissao": doc.get("profissao", " "),
                        "Data Vencimento": doc.get("data_vencimento", " "),
                        "Origem": doc.get("origem", " "),
                        "Codigo Indicacao": doc.get("codigo_indicacao", " "),
                        "Cadastrado Por": doc.get("cadastrado_por", " ")
                    })
                df_risco = pd.DataFrame(dados_risco)
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_risco.to_excel(writer, index=False, sheet_name='Analise de Risco')
                    worksheet = writer.sheets['Analise de Risco']
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                excel_data = output.getvalue()
                st.download_button(label=f"📊 Exportar Análise de Risco ({len(docs_risco)} ativações)", data=excel_data, file_name=f"analise_risco_{inicio.strftime('%Y%m%d')}_{fim.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="export_risco_excel")
            else:
                st.info("ℹ️ Nenhuma ativação no período selecionado")

        st.markdown("---")
        st.subheader("📈 Evolutivo Mensal de Conversão")
        pipeline_mensal = [
            {"$match": {"data_cadastro": {"$gte": inicio.replace(day=1) - relativedelta(months=11), "$lte": fim}}},
            {"$group": {
                "_id": {"ano": {"$year": "$data_cadastro"}, "mes": {"$month": "$data_cadastro"}},
                "total_cadastros": {"$sum": 1},
                "total_ativacoes": {"$sum": {"$cond": [{"$eq": ["$seguiu_ativacao", "Sim"]}, 1, 0]}},
                "nao_ativados_restritivos": {"$sum": {"$cond": [{"$and": [{"$eq": ["$seguiu_ativacao", "Não"]}, {"$eq": ["$restritivo", "Sim"]}]}, 1, 0]}},
                "nao_ativados_outros": {"$sum": {"$cond": [{"$and": [{"$eq": ["$seguiu_ativacao", "Não"]}, {"$ne": ["$restritivo", "Sim"]}]}, 1, 0]}}
            }},
            {"$sort": {"_id.ano": 1, "_id.mes": 1}}
        ]
        dados_mensais = list(clientes_collection.aggregate(pipeline_mensal))
        if dados_mensais:
            df_mensal = pd.DataFrame(dados_mensais)
            df_mensal["mes_ano"] = df_mensal["_id"].apply(lambda x: f"{x['ano']}-{x['mes']:02d}")
            df_mensal["taxa_conversao"] = (df_mensal["total_ativacoes"] / df_mensal["total_cadastros"] * 100).round(2)
            fig_evolucao = go.Figure()
            fig_evolucao.add_trace(go.Bar(x=df_mensal["mes_ano"], y=df_mensal["total_cadastros"], name="Total Cadastros", marker_color="#3498db", text=df_mensal["total_cadastros"], textposition="outside"))
            fig_evolucao.add_trace(go.Bar(x=df_mensal["mes_ano"], y=df_mensal["total_ativacoes"], name="Total Ativações", marker_color="#2ecc71", text=df_mensal["total_ativacoes"], textposition="outside"))
            fig_evolucao.add_trace(go.Scatter(x=df_mensal["mes_ano"], y=df_mensal["taxa_conversao"], name="Taxa Conversão (%)", mode="lines+markers+text", line=dict(color="#e74c3c", width=3), marker=dict(size=8), text=df_mensal["taxa_conversao"].apply(lambda x: f"{x}%"), textposition="top center", yaxis="y2"))
            fig_evolucao.update_layout(title="Evolução Mensal: Cadastros, Ativações e Taxa de Conversão", barmode="group", xaxis=dict(title="Mês/Ano"), yaxis=dict(title="Quantidade", side="left"), yaxis2=dict(title="Taxa Conversão (%)", overlaying="y", side="right", showgrid=False), legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1), height=500, hovermode="x unified")
            st.plotly_chart(fig_evolucao, use_container_width=True)
            with st.expander("📋 Ver Dados Detalhados"):
                df_display = df_mensal[["mes_ano", "total_cadastros", "total_ativacoes", "taxa_conversao"]].copy()
                df_display.columns = ["Mês/Ano", "Total Cadastros", "Total Ativações", "Taxa Conversão (%)"]
                st.dataframe(df_display, use_container_width=True)

        st.markdown("---")
        st.subheader("📊 Análise de Não Conversão")
        analise_data = _analise_nao_conversao_detalhada(clientes_collection, inicio, fim)
        df_detalhado = _tabela_nao_conversao_detalhada(clientes_collection, inicio, fim)
        if analise_data["total_nao_conversao"] > 0:
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Não Convertidos", analise_data["total_nao_conversao"])
            with col2:
                perc_restritivos = (analise_data["restritivos"] / analise_data["total_nao_conversao"] * 100) if analise_data["total_nao_conversao"] > 0 else 0
                st.metric("Por Restrição", analise_data["restritivos"], f"{perc_restritivos:.1f}%")
            with col3:
                perc_outros = (analise_data["outros_motivos"] / analise_data["total_nao_conversao"] * 100) if analise_data["total_nao_conversao"] > 0 else 0
                st.metric("Outros Motivos", analise_data["outros_motivos"], f"{perc_outros:.1f}%")
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                motivos = list(analise_data["detalhes_por_motivo"].keys())
                valores = [analise_data["detalhes_por_motivo"][m]["total"] for m in motivos]
                fig_pizza = go.Figure(data=[go.Pie(labels=motivos, values=valores, hole=0.3, textinfo='label+percent', textposition='outside', automargin=True)])
                fig_pizza.update_layout(title="Distribuição por Motivo de Recusa", height=400, showlegend=True, legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1))
                st.plotly_chart(fig_pizza, use_container_width=True)
            with col2:
                df_comp = pd.DataFrame({"Categoria": ["Restritivos", "Outros Motivos"], "Quantidade": [analise_data["restritivos"], analise_data["outros_motivos"]]})
                fig_barras = px.bar(df_comp, x="Categoria", y="Quantidade", color="Categoria", color_discrete_map={"Restritivos": "#e74c3c", "Outros Motivos": "#f39c12"}, text="Quantidade")
                fig_barras.update_layout(title="Comparativo: Restritivos vs Outros Motivos", height=400, showlegend=False)
                st.plotly_chart(fig_barras, use_container_width=True)
            st.markdown("### 📊 Top Motivos de Recusa")
            df_motivos = pd.DataFrame([{"Motivo": motivo, "Total": data["total"]} for motivo, data in analise_data["detalhes_por_motivo"].items()]).sort_values("Total", ascending=True)
            if not df_motivos.empty:
                fig_top = px.bar(df_motivos, y="Motivo", x="Total", orientation='h', color="Total", color_continuous_scale="Reds", text="Total")
                fig_top.update_layout(height=max(300, len(df_motivos) * 40), showlegend=False, xaxis_title="Quantidade", yaxis_title="")
                st.plotly_chart(fig_top, use_container_width=True)
            st.markdown("### 💡 Insights")
            col1, col2 = st.columns(2)
            with col1:
                if analise_data["restritivos"] > analise_data["outros_motivos"]:
                    st.warning(f"⚠️ **Maioria por restrição**: {perc_restritivos:.1f}% das não conversões são por restrições financeiras")
                else:
                    st.info(f"ℹ️ **Maioria por outros motivos**: {perc_outros:.1f}% das não conversões são por outros fatores")
            with col2:
                motivo_top = df_motivos.iloc[-1]["Motivo"] if not df_motivos.empty else "N/A"
                total_top = df_motivos.iloc[-1]["Total"] if not df_motivos.empty else 0
                st.metric("Motivo Mais Comum", motivo_top, f"{total_top} ocorrências")
            st.markdown("---")
            st.markdown("### 📋 Detalhamento das Não Conversões")
            if not df_detalhado.empty:
                col1, col2, col3 = st.columns(3)
                with col1:
                    filtro_categoria = st.multiselect("Filtrar por Categoria", options=df_detalhado["Categoria"].unique(), default=df_detalhado["Categoria"].unique())
                with col2:
                    filtro_motivo = st.multiselect("Filtrar por Motivo", options=df_detalhado["Motivo Recusa"].unique(), default=df_detalhado["Motivo Recusa"].unique())
                with col3:
                    filtro_restritivo = st.selectbox("É Restritivo?", options=["Todos", "Sim", "Não"], index=0)
                df_filtrado = df_detalhado.copy()
                if filtro_categoria:
                    df_filtrado = df_filtrado[df_filtrado["Categoria"].isin(filtro_categoria)]
                if filtro_motivo:
                    df_filtrado = df_filtrado[df_filtrado["Motivo Recusa"].isin(filtro_motivo)]
                if filtro_restritivo != "Todos":
                    df_filtrado = df_filtrado[df_filtrado["É Restritivo?"] == filtro_restritivo]
                st.dataframe(df_filtrado, use_container_width=True, hide_index=True)
                st.markdown("#### 📥 Exportar Dados")
                col1, col2 = st.columns(2)
                with col1:
                    csv = df_filtrado.to_csv(index=False, sep=';').encode('utf-8')
                    st.download_button(label="📥 Exportar CSV", data=csv, file_name=f"nao_conversao_{inicio.strftime('%Y%m%d')}_{fim.strftime('%Y%m%d')}.csv", mime="text/csv")
                with col2:
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_filtrado.to_excel(writer, index=False, sheet_name='Não Conversão')
                        worksheet = writer.sheets['Não Conversão']
                        header_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    excel_data = output.getvalue()
                    st.download_button(label="📊 Exportar Excel", data=excel_data, file_name=f"nao_conversao_{inicio.strftime('%Y%m%d')}_{fim.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.markdown("---")
            st.markdown("### 📈 Evolução Temporal por Motivo")
            df_evolucao = _evolucao_motivos(clientes_collection, inicio, fim)
            if not df_evolucao.empty:
                df_pivot = df_evolucao.pivot_table(index="mes_ano", columns="motivo", values="total", fill_value=0).reset_index()
                df_melt = df_pivot.melt(id_vars=["mes_ano"], var_name="Motivo", value_name="Quantidade")
                fig_evolucao = px.line(df_melt, x="mes_ano", y="Quantidade", color="Motivo", markers=True, title="Evolução Mensal dos Motivos de Recusa")
                fig_evolucao.update_layout(height=400, xaxis_title="Mês/Ano", yaxis_title="Quantidade", hovermode="x unified")
                st.plotly_chart(fig_evolucao, use_container_width=True)
        else:
            st.success("✅ Nenhuma não conversão no período selecionado (taxa de conversão 100%!)")

        st.markdown("---")
        st.subheader("📋 Resumo Executivo")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("### 🎯 Principais Insights")
            st.markdown(f"""
            - **Total de cadastros:** {conversao_data['total_cadastros']}
            - **Taxa de conversão geral:** {conversao_data['taxa_geral']}%
            - **Clientes com restrição:** {conversao_data['restritivos']} ({conversao_data['taxa_restritivos']}% converteram)
            - **Média de tempo para ativação:** {tempo_data['media_dias']} dias
            - **Indicações que ativaram:** {df_indicacoes_mes['total'].sum() if not df_indicacoes_mes.empty else 0}
            """)
        with col2:
            st.markdown("### ⚠️ Pontos de Atenção")
            if conversao_data['taxa_restritivos'] < conversao_data['taxa_nao_restritivos']:
                st.warning(f"Taxa de conversão com restrição ({conversao_data['taxa_restritivos']}%) é menor que sem restrição ({conversao_data['taxa_nao_restritivos']})")
            if tempo_data['media_dias'] > 7:
                st.warning(f"Tempo médio para ativação ({tempo_data['media_dias']} dias) está acima de 7 dias")
            if restricoes_data['media_restricoes'] != "N/A" and restricoes_data['media_restricoes'] > 2:
                st.warning(f"Média de restrições por cliente ({restricoes_data['media_restricoes']}) está elevada")
